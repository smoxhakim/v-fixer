import csv
import io
import re
import uuid
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from xml.etree import ElementTree

from django.conf import settings
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook
from PIL import Image

from .models import Category, Product

MAX_PRODUCT_ROWS = 5000
MAX_IMAGES_PER_ROW = 5
MAX_IMAGE_SIZE_BYTES = 2 * 1024 * 1024
ALLOWED_XLSX_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
    "application/zip",
}
ALLOWED_ZIP_CONTENT_TYPES = {
    "application/zip",
    "application/x-zip-compressed",
    "multipart/x-zip",
    "application/octet-stream",
}
OUT_OF_STOCK_MARKERS = {"out stock", "out", "outofstock", "0", "no stock"}
PRODUCT_HEADERS = ("IMAGE", "REF", "PRICE", "PRICE ACHA", "QT", "CATEGORY")


@dataclass
class RowError:
    row: int
    message: str

    def to_dict(self):
        return {"row": self.row, "message": self.message}


def _normalize_slug(raw: str) -> str:
    return slugify((raw or "").strip())


def _normalize_price_text(raw_value) -> str:
    text = "" if raw_value is None else str(raw_value).strip()
    text = text.replace(" ", "")
    if "." in text:
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    return text


def parse_first_decimal(raw_value):
    text = _normalize_price_text(raw_value)
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def parse_stock(raw_value):
    text = "" if raw_value is None else str(raw_value).strip()
    if not text:
        return 0
    compact = text.lower().replace(" ", "")
    if compact == "outstock":
        compact = "outofstock"
    if compact in OUT_OF_STOCK_MARKERS:
        return 0
    num_match = re.search(r"-?\d+", text)
    if not num_match:
        return None
    try:
        return int(num_match.group(0))
    except ValueError:
        return None


def _normalize_zip_path(base_dir: str, target: str):
    target = str(target).replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    base_parts = [p for p in base_dir.split("/") if p]
    for part in target.split("/"):
        if not part or part == ".":
            continue
        if part == "..":
            if base_parts:
                base_parts.pop()
            continue
        base_parts.append(part)
    return "/".join(base_parts)


def _build_unique_slug(base_slug: str, used_slugs: set[str]) -> str:
    base = base_slug or "product"
    if base not in used_slugs:
        used_slugs.add(base)
        return base
    i = 2
    while True:
        candidate = f"{base}-{i}"
        if candidate not in used_slugs:
            used_slugs.add(candidate)
            return candidate
        i += 1


def _validate_xlsx_upload(upload):
    suffix = Path(upload.name or "").suffix.lower()
    content_type = (getattr(upload, "content_type", "") or "").lower()
    if suffix == ".xlsm":
        return False, "Macro-enabled Excel (.xlsm) is not allowed."
    if suffix != ".xlsx":
        return False, "Only .xlsx files are supported."
    if content_type and content_type not in ALLOWED_XLSX_CONTENT_TYPES:
        return False, f"Unsupported content type: {content_type}"
    return True, ""


def _validate_zip_upload(upload):
    suffix = Path(upload.name or "").suffix.lower()
    content_type = (getattr(upload, "content_type", "") or "").lower()
    if suffix != ".zip":
        return False, "Only .zip files are supported for archive import."
    if content_type and content_type not in ALLOWED_ZIP_CONTENT_TYPES:
        return False, f"Unsupported archive content type: {content_type}"
    return True, ""


def import_categories_csv(upload):
    content = upload.read()
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return {"created": 0, "updated": 0, "errors": [{"row": 1, "message": "CSV must be UTF-8."}]}

    reader = csv.DictReader(io.StringIO(decoded))
    required = {"name", "slug"}
    if not reader.fieldnames:
        return {"created": 0, "updated": 0, "errors": [{"row": 1, "message": "CSV has no headers."}]}
    missing = required - set(map(str.strip, reader.fieldnames))
    if missing:
        return {
            "created": 0,
            "updated": 0,
            "errors": [{"row": 1, "message": f"Missing required columns: {', '.join(sorted(missing))}"}],
        }

    errors = []
    parsed_rows = []
    seen_slugs = set()
    for idx, row in enumerate(reader, start=2):
        name = (row.get("name") or "").strip()
        row_slug = _normalize_slug(row.get("slug") or "")
        icon = (row.get("icon") or "").strip() or None
        parent_slug = _normalize_slug(row.get("parent_slug") or "")
        parent_slug = parent_slug or None
        if not name or not row_slug:
            errors.append(RowError(idx, "name and slug are required.").to_dict())
            continue
        if row_slug in seen_slugs:
            errors.append(RowError(idx, f"Duplicate slug in file: {row_slug}").to_dict())
            continue
        seen_slugs.add(row_slug)
        parsed_rows.append(
            {
                "row": idx,
                "name": name,
                "slug": row_slug,
                "icon": icon,
                "parent_slug": parent_slug,
            }
        )

    created = 0
    updated = 0
    category_map = {c.slug: c for c in Category.objects.all()}
    for entry in parsed_rows:
        with transaction.atomic():
            category = category_map.get(entry["slug"])
            if category:
                category.name = entry["name"]
                category.icon = entry["icon"]
                category.parent = None
                category.save(update_fields=["name", "icon", "parent"])
                updated += 1
            else:
                category = Category.objects.create(
                    name=entry["name"],
                    slug=entry["slug"],
                    icon=entry["icon"],
                )
                created += 1
            category_map[entry["slug"]] = category

    unresolved = parsed_rows.copy()
    safety = 0
    while unresolved and safety < len(unresolved) + 5:
        safety += 1
        next_round = []
        progressed = False
        for entry in unresolved:
            slug = entry["slug"]
            parent_slug = entry["parent_slug"]
            if not parent_slug:
                continue
            category = category_map.get(slug)
            parent = category_map.get(parent_slug)
            if parent is None:
                next_round.append(entry)
                continue
            if parent.slug == category.slug:
                errors.append(RowError(entry["row"], "Category cannot be its own parent.").to_dict())
                continue
            chain = set()
            cursor = parent
            cyclic = False
            while cursor:
                if cursor.slug == category.slug or cursor.slug in chain:
                    cyclic = True
                    break
                chain.add(cursor.slug)
                cursor = cursor.parent
            if cyclic:
                errors.append(RowError(entry["row"], f"Cycle detected for slug: {slug}").to_dict())
                continue
            if category.parent_id != parent.id:
                category.parent = parent
                category.save(update_fields=["parent"])
                progressed = True
        if not next_round:
            break
        if not progressed and len(next_round) == len(unresolved):
            for entry in next_round:
                errors.append(
                    RowError(entry["row"], f"Unknown parent_slug: {entry['parent_slug']}").to_dict()
                )
            break
        unresolved = next_round

    return {"created": created, "updated": updated, "errors": errors}


def _read_upload_bytes(upload):
    if hasattr(upload, "seek"):
        upload.seek(0)
    return upload.read()


def _sort_row_image_map(row_map):
    for row, items in row_map.items():
        row_map[row] = sorted(items, key=lambda item: item["col"])
    return row_map


def _merge_image_row_maps(primary: dict, secondary: dict) -> dict:
    """Prefer primary (zip/XML); add secondary (openpyxl) cells not already present on that row."""
    out = {r: list(items) for r, items in primary.items()}
    for row, items in secondary.items():
        cols = {i["col"] for i in out.get(row, [])}
        for item in items:
            if item["col"] not in cols:
                out.setdefault(row, []).append(item)
                cols.add(item["col"])
    return _sort_row_image_map(out)


def _image_row_map_from_openpyxl(data: bytes):
    row_map = {}
    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
        ws = workbook.active
        images = getattr(ws, "_images", [])
        for img in images:
            anchor = getattr(img, "anchor", None)
            marker = getattr(anchor, "_from", None)
            if marker is None:
                continue
            row_1_based = int(getattr(marker, "row", 0)) + 1
            col_1_based = int(getattr(marker, "col", 0)) + 1
            raw_bytes = None
            try:
                raw_bytes = img._data()
            except Exception:
                raw_bytes = None
            if not raw_bytes:
                continue
            ext = ".png"
            row_map.setdefault(row_1_based, []).append({"bytes": raw_bytes, "ext": ext, "col": col_1_based})
    except Exception:
        return {}
    return _sort_row_image_map(row_map)


def _image_row_map_from_zip(data: bytes):
    row_map = {}
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            ns_wb = {
                "w": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
            }
            workbook_root = ElementTree.fromstring(zf.read("xl/workbook.xml"))
            sheets = workbook_root.findall("w:sheets/w:sheet", ns_wb)
            if not sheets:
                return row_map
            active_idx = 0
            wb_view = workbook_root.find("w:bookViews/w:workbookView", ns_wb)
            if wb_view is not None and wb_view.attrib.get("activeTab") is not None:
                try:
                    active_idx = int(wb_view.attrib["activeTab"])
                except ValueError:
                    active_idx = 0
            active_idx = max(0, min(active_idx, len(sheets) - 1))
            sheet_el = sheets[active_idx]
            sheet_rid = sheet_el.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            workbook_rels_root = ElementTree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            sheet_target = None
            for rel in workbook_rels_root.findall("pr:Relationship", ns_wb):
                if rel.attrib.get("Id") == sheet_rid:
                    sheet_target = rel.attrib.get("Target")
                    break
            if not sheet_target:
                return row_map
            sheet_path = _normalize_zip_path("xl", sheet_target)
            sheet_rel_path = f"xl/worksheets/_rels/{Path(sheet_path).name}.rels"
            if sheet_rel_path not in zf.namelist():
                return row_map
            ns_rel = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rel_root = ElementTree.fromstring(zf.read(sheet_rel_path))
            drawing_target = None
            for rel in rel_root.findall("r:Relationship", ns_rel):
                r_type = rel.attrib.get("Type", "")
                if r_type.endswith("/drawing"):
                    drawing_target = rel.attrib.get("Target")
                    break
            if not drawing_target:
                return row_map
            drawing_path = _normalize_zip_path("xl/worksheets", drawing_target)
            if drawing_path not in zf.namelist():
                drawing_path = f"xl/drawings/{Path(drawing_target).name}"
                if drawing_path not in zf.namelist():
                    return row_map
            drawing_rels_path = f"xl/drawings/_rels/{Path(drawing_path).name}.rels"
            if drawing_rels_path not in zf.namelist():
                return row_map

            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
            }
            rels_root = ElementTree.fromstring(zf.read(drawing_rels_path))
            media_by_rid = {}
            for rel in rels_root.findall("pr:Relationship", ns):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target", "")
                media_path = _normalize_zip_path("xl/drawings", target)
                if media_path not in zf.namelist():
                    media_path = f"xl/media/{Path(target).name}"
                media_by_rid[rid] = media_path

            drawing_root = ElementTree.fromstring(zf.read(drawing_path))
            anchors = drawing_root.findall("xdr:twoCellAnchor", ns) + drawing_root.findall("xdr:oneCellAnchor", ns)
            for anchor in anchors:
                row_node = anchor.find("xdr:from/xdr:row", ns)
                col_node = anchor.find("xdr:from/xdr:col", ns)
                blip = anchor.find(".//a:blip", ns)
                if row_node is None or blip is None:
                    continue
                rid = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                media_path = media_by_rid.get(rid)
                if not media_path or media_path not in zf.namelist():
                    continue
                row_1_based = int(row_node.text or "0") + 1
                col_1_based = int(col_node.text or "0") + 1 if col_node is not None else 1
                ext = Path(media_path).suffix.lower() or ".jpg"
                row_map.setdefault(row_1_based, []).append(
                    {"bytes": zf.read(media_path), "ext": ext, "col": col_1_based}
                )
    except zipfile.BadZipFile:
        return {}
    return _sort_row_image_map(row_map)


def _extract_row_images_from_xlsx(data: bytes):
    from_zip = _image_row_map_from_zip(data)
    from_oox = _image_row_map_from_openpyxl(data)
    merged = _merge_image_row_maps(from_zip, from_oox)
    return merged


def _process_image_bytes(image_bytes: bytes, output_hint: str = ".webp"):
    with Image.open(io.BytesIO(image_bytes)) as img:
        img = img.convert("RGB")
        max_side = 1920
        img.thumbnail((max_side, max_side))
        fmt = "WEBP" if output_hint == ".webp" else "JPEG"
        ext = ".webp" if fmt == "WEBP" else ".jpg"
        out = io.BytesIO()
        if fmt == "WEBP":
            img.save(out, format=fmt, quality=82, method=6)
        else:
            img.save(out, format=fmt, quality=84, optimize=True)
        data = out.getvalue()
        if len(data) > MAX_IMAGE_SIZE_BYTES:
            out = io.BytesIO()
            if fmt == "WEBP":
                img.save(out, format=fmt, quality=70, method=4)
            else:
                img.save(out, format=fmt, quality=70, optimize=True)
            data = out.getvalue()
        if len(data) > MAX_IMAGE_SIZE_BYTES:
            raise ValueError("Image exceeds 2MB after compression.")
        return data, ext


def _save_image_for_product(image_bytes: bytes, slug: str, idx: int, import_id: str):
    processed, ext = _process_image_bytes(image_bytes)
    rel_dir = Path("imports") / import_id
    abs_dir = Path(settings.MEDIA_ROOT) / rel_dir
    abs_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{slug}-{idx}{ext}"
    abs_path = abs_dir / filename
    abs_path.write_bytes(processed)
    media_url = settings.MEDIA_URL.rstrip("/")
    return f"{media_url}/{rel_dir.as_posix()}/{filename}"


def _parse_product_rows_from_bytes(data: bytes):
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    ws = workbook.active

    header_map = {}
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value is None:
            continue
        header_map[str(header_value).strip().upper()] = col

    missing = [h for h in PRODUCT_HEADERS if h not in header_map]
    if missing:
        return None, None, [RowError(1, f"Missing required headers: {', '.join(missing)}").to_dict()]

    row_images = _extract_row_images_from_xlsx(data)
    parsed = []
    errors = []
    for row_idx in range(2, ws.max_row + 1):
        if row_idx > MAX_PRODUCT_ROWS + 1:
            errors.append(RowError(row_idx, f"Exceeded max rows ({MAX_PRODUCT_ROWS}).").to_dict())
            break
        ref = ws.cell(row=row_idx, column=header_map["REF"]).value
        price_raw = ws.cell(row=row_idx, column=header_map["PRICE"]).value
        cost_raw = ws.cell(row=row_idx, column=header_map["PRICE ACHA"]).value
        qt_raw = ws.cell(row=row_idx, column=header_map["QT"]).value
        category_raw = ws.cell(row=row_idx, column=header_map["CATEGORY"]).value

        if all(v in (None, "") for v in [ref, price_raw, cost_raw, qt_raw, category_raw]):
            continue

        ref_text = ("" if ref is None else str(ref)).strip()
        if not ref_text:
            errors.append(RowError(row_idx, "REF is required.").to_dict())
            continue
        price = parse_first_decimal(price_raw)
        if price is None:
            errors.append(RowError(row_idx, "PRICE has no valid number.").to_dict())
            continue
        cost_price = parse_first_decimal(cost_raw)
        stock = parse_stock(qt_raw)
        if stock is None:
            errors.append(RowError(row_idx, "QT has no valid stock number.").to_dict())
            continue
        category_slug = _normalize_slug("" if category_raw is None else str(category_raw))
        category_slug = category_slug or None

        images = row_images.get(row_idx, [])
        parsed.append(
            {
                "row": row_idx,
                "ref": ref_text,
                "price": price,
                "cost_price": cost_price,
                "stock": stock,
                "category_slug": category_slug,
                "images": images,
            }
        )

    return parsed, data, errors


def _parse_product_rows(upload):
    data = _read_upload_bytes(upload)
    return _parse_product_rows_from_bytes(data)


def preview_products_xlsx(upload):
    ok, msg = _validate_xlsx_upload(upload)
    if not ok:
        return {"created": 0, "errors": [{"row": 1, "message": msg}], "rows": []}

    data = _read_upload_bytes(upload)
    rows, _data, errors = _parse_product_rows_from_bytes(data)
    if rows is None:
        return {"created": 0, "errors": errors, "rows": []}

    categories_map = {c.slug: c for c in Category.objects.all()}
    preview_rows = []
    for row in rows:
        slug = row["category_slug"]
        if slug is None:
            resolved = None
            category_resolved = True
        else:
            resolved = categories_map.get(slug)
            category_resolved = bool(resolved)
        preview_rows.append(
            {
                "row": row["row"],
                "ref": row["ref"],
                "price": str(row["price"]),
                "costPrice": str(row["cost_price"]) if row["cost_price"] is not None else None,
                "stock": row["stock"],
                "categorySlug": slug,
                "categoryResolved": category_resolved,
                "imagesDetected": min(len(row["images"]), MAX_IMAGES_PER_ROW),
            }
        )
        if slug is not None and not resolved:
            errors.append(RowError(row["row"], f"Unknown category slug: {slug}").to_dict())

    return {"created": 0, "errors": errors, "rows": preview_rows}


def import_products_xlsx(upload):
    ok, msg = _validate_xlsx_upload(upload)
    if not ok:
        return {"created": 0, "errors": [{"row": 1, "message": msg}]}

    data = _read_upload_bytes(upload)
    rows, _data, errors = _parse_product_rows_from_bytes(data)
    if rows is None:
        return {"created": 0, "errors": errors}

    categories_map = {c.slug: c for c in Category.objects.all()}
    used_slugs = set(Product.objects.values_list("slug", flat=True))
    created = 0
    import_id = str(uuid.uuid4())

    for row in rows:
        slug = row["category_slug"]
        if slug is None:
            category = None
        else:
            category = categories_map.get(slug)
            if category is None:
                errors.append(RowError(row["row"], f"Unknown category slug: {slug}").to_dict())
                continue

        base_slug = slugify(row["ref"])
        slug = _build_unique_slug(base_slug, used_slugs)
        image_urls = []
        row_images = row["images"][:MAX_IMAGES_PER_ROW]
        if len(row["images"]) > MAX_IMAGES_PER_ROW:
            errors.append(RowError(row["row"], f"More than {MAX_IMAGES_PER_ROW} images; extra images ignored.").to_dict())
        for idx, image_data in enumerate(row_images, start=1):
            try:
                image_urls.append(_save_image_for_product(image_data["bytes"], slug, idx, import_id))
            except Exception as exc:
                errors.append(RowError(row["row"], f"Image skipped: {exc}").to_dict())

        with transaction.atomic():
            Product.objects.create(
                name=row["ref"],
                slug=slug,
                category=category,
                price=row["price"],
                cost_price=row["cost_price"],
                stock=max(0, row["stock"]),
                images=image_urls,
                rating=0,
            )
            created += 1

    return {"created": created, "errors": errors}


def import_products_zip(upload):
    ok, msg = _validate_zip_upload(upload)
    if not ok:
        return {"created": 0, "errors": [{"row": 1, "message": msg}]}

    data = upload.read()
    errors = []
    created = 0
    import_id = str(uuid.uuid4())
    categories_map = {c.slug: c for c in Category.objects.all()}
    used_slugs = set(Product.objects.values_list("slug", flat=True))

    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        return {"created": 0, "errors": [{"row": 1, "message": "Invalid ZIP archive."}]}

    if "products.csv" not in zf.namelist():
        return {"created": 0, "errors": [{"row": 1, "message": "Archive must include products.csv."}]}

    csv_bytes = zf.read("products.csv")
    try:
        csv_text = csv_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        return {"created": 0, "errors": [{"row": 1, "message": "products.csv must be UTF-8."}]}

    reader = csv.DictReader(io.StringIO(csv_text))
    for row_idx, row in enumerate(reader, start=2):
        ref = (row.get("REF") or row.get("ref") or "").strip()
        price = parse_first_decimal(row.get("PRICE") or row.get("price"))
        cost_price = parse_first_decimal(row.get("PRICE ACHA") or row.get("price_acha"))
        stock = parse_stock(row.get("QT") or row.get("qt"))
        category_slug = _normalize_slug(row.get("CATEGORY") or row.get("category") or "")
        category_slug = category_slug or None
        image_paths = (row.get("image_paths") or "").strip()

        if not ref or price is None or stock is None:
            errors.append(RowError(row_idx, "Invalid row data in products.csv.").to_dict())
            continue
        if not category_slug:
            category = None
        else:
            category = categories_map.get(category_slug)
            if category is None:
                errors.append(RowError(row_idx, f"Unknown category slug: {category_slug}").to_dict())
                continue
        slug = _build_unique_slug(slugify(ref), used_slugs)
        image_urls = []
        if image_paths:
            for idx, rel_path in enumerate(image_paths.split("|"), start=1):
                rel_path = rel_path.strip().lstrip("/")
                if not rel_path:
                    continue
                if idx > MAX_IMAGES_PER_ROW:
                    errors.append(RowError(row_idx, f"More than {MAX_IMAGES_PER_ROW} images; extra ignored.").to_dict())
                    break
                if rel_path not in zf.namelist():
                    errors.append(RowError(row_idx, f"Missing image in archive: {rel_path}").to_dict())
                    continue
                try:
                    image_urls.append(_save_image_for_product(zf.read(rel_path), slug, idx, import_id))
                except Exception as exc:
                    errors.append(RowError(row_idx, f"Image skipped: {exc}").to_dict())

        with transaction.atomic():
            Product.objects.create(
                name=ref,
                slug=slug,
                category=category,
                price=price,
                cost_price=cost_price,
                stock=max(0, stock),
                images=image_urls,
                rating=0,
            )
            created += 1

    return {"created": created, "errors": errors}

